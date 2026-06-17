"""Export comment records to Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from bilibili_client import CommentRecord


def export_comments_to_excel(
    comments: list[CommentRecord],
    output_path: str | Path,
    summary: dict,
) -> Path:
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "用户评论"

    headers = [
        "序号",
        "评论类型",
        "用户名",
        "UID",
        "评论内容",
        "点赞数",
        "发布时间",
        "评论ID",
        "根评论ID",
        "回复对象",
    ]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, comment in enumerate(comments, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=comment.comment_type)
        ws.cell(row=row_idx, column=3, value=comment.username)
        ws.cell(row=row_idx, column=4, value=comment.uid)
        ws.cell(row=row_idx, column=5, value=comment.content)
        ws.cell(row=row_idx, column=6, value=comment.likes)
        ws.cell(row=row_idx, column=7, value=comment.publish_time)
        ws.cell(row=row_idx, column=8, value=comment.rpid)
        ws.cell(row=row_idx, column=9, value=comment.root_rpid or "")
        ws.cell(row=row_idx, column=10, value=comment.reply_to)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    column_widths = [8, 12, 18, 14, 60, 10, 20, 14, 14, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    info = wb.create_sheet("视频信息")
    info_rows = [
        ("视频标题", summary.get("title", "")),
        ("BV号", summary.get("bvid", "")),
        ("AID", summary.get("aid", "")),
        ("目标用户", summary.get("username", "")),
        ("主评论总数", summary.get("main_total", 0)),
        ("子评论扫描数", summary.get("sub_total", 0)),
        ("匹配评论数", summary.get("matched_total", 0)),
    ]
    for row_idx, (label, value) in enumerate(info_rows, start=1):
        info.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        info.cell(row=row_idx, column=2, value=value)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 50

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
